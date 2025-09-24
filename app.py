import os
import io
import re
import string
import hashlib
from datetime import datetime

import pandas as pd
import streamlit as st
from dotenv import load_dotenv

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import sqlalchemy as sa
from sqlalchemy import text

from openai import AzureOpenAI

# -------------------------
# Load environment
# -------------------------
load_dotenv()

try:
    AZURE_OPENAI_ENDPOINT = st.secrets["AZURE_OPENAI_ENDPOINT"]
    AZURE_OPENAI_API_KEY = st.secrets["AZURE_OPENAI_API_KEY"]
    AZURE_OPENAI_DEPLOYMENT = st.secrets["AZURE_OPENAI_DEPLOYMENT"]
    DATABASE = st.secrets["DATABASE"]
except Exception:
    AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
    AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
    AZURE_OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")
    DATABASE = os.getenv("DATABASE")

if not AZURE_OPENAI_API_KEY or not AZURE_OPENAI_ENDPOINT or not AZURE_OPENAI_DEPLOYMENT:
    st.error("Missing Azure OpenAI credentials. Set AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_API_KEY, AZURE_OPENAI_DEPLOYMENT.")
    st.stop()

DATABASE_URL = (DATABASE or "").strip()

MASTER_XLSX = "genai_job_impact_master.xlsx"
ALL_JOBS_SHEET = "All Jobs"
SYNTHESIS_SHEET = "Synthesis"

st.set_page_config(page_title="GenAI Job Impact Analyst — Postgres (Render)", layout="wide")
st.title("💼 GenAI Job Impact Analyst — Postgres (Render)")

# -------------------------
# Initialize Azure OpenAI client
# -------------------------
client = AzureOpenAI(
    api_key=AZURE_OPENAI_API_KEY,
    api_version="2024-06-01",
    azure_endpoint=AZURE_OPENAI_ENDPOINT,
)

# -------------------------
# Helper functions (verbose)
# -------------------------
def normalize_task(task: str) -> str:
    if not isinstance(task, str):
        task = "" if pd.isna(task) else str(task)
    t = task.lower().strip()
    t = re.sub(r'\s+', ' ', t)
    table = str.maketrans('', '', string.punctuation)
    t = t.translate(table)
    return t

def extract_roles_from_text(raw_text: str) -> list[str]:
    blocks = re.split(r'\n\s*---\s*\n', raw_text.strip(), flags=re.MULTILINE)
    return [b.strip() for b in blocks if b.strip()]

def role_name_from_jobdesc(jd: str, index: int) -> str:
    m = re.search(r'(?i)^ *Job Title:\s*(.+)$', jd, flags=re.MULTILINE)
    if m:
        name = m.group(1).strip()
    else:
        first = next((ln.strip() for ln in jd.splitlines() if ln.strip()), f"Job_{index+1}")
        name = first.split("|")[0].split(" - ")[0].strip()
    name = re.sub(r'[\[\]\*\?/\\:]', "", name)[:120]
    return name or f"Job_{index+1}"

def parse_markdown_table(md_text: str) -> pd.DataFrame:
    lines = [ln.rstrip() for ln in md_text.splitlines()]
    table_lines = [ln for ln in lines if "|" in ln]
    if not table_lines:
        return pd.DataFrame()
    sep_pat = r'^\s*\|?\s*[-:]+(?:\s*\|\s*[-:]+)*\s*\|?\s*$'
    table_lines = [ln for ln in table_lines if not re.match(sep_pat, ln)]
    if not table_lines:
        return pd.DataFrame()
    rows = []
    for ln in table_lines:
        parts = [cell.strip() for cell in ln.strip().strip("|").split("|")]
        rows.append(parts)
    header = rows[0]
    data = rows[1:] if len(rows) > 1 else []
    while header and header[-1] == "":
        header = header[:-1]
        data = [r[:-1] for r in data]
    try:
        df = pd.DataFrame(data, columns=header)
    except Exception:
        max_cols = max(len(r) for r in rows)
        cols = [f"col_{i}" for i in range(max_cols)]
        df = pd.DataFrame([r + [""]*(max_cols-len(r)) for r in rows[1:]], columns=cols)
    df = df.loc[:, ~(df.columns.str.strip() == "")]
    return df

def split_table_and_synthesis(text: str) -> tuple[str, str]:
    parts = text.split("Synthesis:")
    if len(parts) == 2:
        return parts[0], parts[1].strip()
    parts = re.split(r'(?i)Synthèse\s*:', text)
    if len(parts) == 2:
        return parts[0], parts[1].strip()
    return text, ""

# Excel helpers
def load_master_excel() -> tuple[pd.DataFrame, pd.DataFrame]:
    if not os.path.exists(MASTER_XLSX):
        return pd.DataFrame(), pd.DataFrame(columns=["Job Title", "Synthesis", "Run ID", "JD Hash"])
    try:
        xl = pd.ExcelFile(MASTER_XLSX)
        all_jobs = pd.read_excel(xl, sheet_name=ALL_JOBS_SHEET) if ALL_JOBS_SHEET in xl.sheet_names else pd.DataFrame()
        syn = pd.read_excel(xl, sheet_name=SYNTHESIS_SHEET) if SYNTHESIS_SHEET in xl.sheet_names else pd.DataFrame(columns=["Job Title", "Synthesis", "Run ID", "JD Hash"])
        return all_jobs, syn
    except Exception as e:
        st.warning(f"Failed to read master Excel: {e}")
        return pd.DataFrame(), pd.DataFrame(columns=["Job Title", "Synthesis", "Run ID", "JD Hash"])

def write_master_excel(all_jobs: pd.DataFrame, synthesis: pd.DataFrame) -> io.BytesIO:
    if "Job Title" in all_jobs.columns:
        cols = all_jobs.columns.tolist()
        cols = ["Job Title"] + [c for c in cols if c != "Job Title"]
        all_jobs = all_jobs[cols]
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        all_jobs.to_excel(writer, sheet_name=ALL_JOBS_SHEET, index=False)
        synthesis.to_excel(writer, sheet_name=SYNTHESIS_SHEET, index=False)
    wb = load_workbook(MASTER_XLSX)
    if ALL_JOBS_SHEET in wb.sheetnames:
        ws = wb[ALL_JOBS_SHEET]
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for c in col:
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if c.value:
                    l = len(str(c.value))
                    if l > max_len:
                        max_len = l
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
    wb.save(MASTER_XLSX)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# Postgres helpers (robust upsert)
def get_engine() -> sa.Engine:
    return sa.create_engine(DATABASE_URL, pool_pre_ping=True, future=True)

def ensure_sql_schema(engine: sa.Engine):
    with engine.begin() as conn:
        conn.execute(text("""
        CREATE TABLE IF NOT EXISTS all_jobs (
            id SERIAL PRIMARY KEY,
            job_title TEXT,
            task TEXT,
            time_allocation TEXT,
            ai_impact_score TEXT,
            impact_explanation TEXT,
            task_transformation TEXT,
            tooling_nature TEXT,
            job_category TEXT,
            Automation_Solution TEXT,
            run_id TEXT,
            jd_hash TEXT,
            task_norm TEXT,
            CONSTRAINT uq_job_task UNIQUE (job_title, task_norm)
        );
        """))
        conn.execute(text("""
        CREATE TABLE IF NOT EXISTS synthesis (
            id SERIAL PRIMARY KEY,
            job_title TEXT,
            synthesis TEXT,
            run_id TEXT,
            jd_hash TEXT,
            CONSTRAINT uq_job_synthesis UNIQUE (job_title, jd_hash)
        );
        """))

def upsert_all_jobs_sql(engine: sa.Engine, df: pd.DataFrame):
    if df is None or df.empty:
        return
    expected_cols = [
        "Job Title",
        "Task",
        "Time allocation %",
        "AI Impact Score (0–100)",
        "Impact Explanation",
        "Task Transformation %",
        "Tooling nature % generic vs specific",
        "Job Category",
        "Automation Solution",
        "Run ID",
        "JD Hash",
        "task_norm"
    ]
    col_map = {}
    for c in df.columns:
        cc = c.strip()
        lower = cc.lower()
        if lower in ["task", "tasks"]:
            col_map[c] = "Task"
        elif "job category" in lower:
            col_map[c] = "Job Category"
        elif "time" in lower and "alloc" in lower:
            col_map[c] = "Time allocation %"
        elif "ai impact" in lower or "impact score" in lower:
            col_map[c] = "AI Impact Score (0–100)"
        elif "impact explanation" in lower or ("explanation" in lower and "impact" in lower):
            col_map[c] = "Impact Explanation"
        elif "task transformation" in lower or "transformation" in lower:
            col_map[c] = "Task Transformation %"
        elif "tooling" in lower:
            col_map[c] = "Tooling nature % generic vs specific"
        elif "automation solution" in lower or ("solution" in lower and "automation" in lower):
            col_map[c] = "Automation Solution"
        elif "job title" in lower or lower == "title":
            col_map[c] = "Job Title"
        elif lower in ["run id", "run_id", "runid"]:
            col_map[c] = "Run ID"
        elif lower in ["jd hash", "jd_hash", "jdhash"]:
            col_map[c] = "JD Hash"
        elif lower == "task_norm":
            col_map[c] = "task_norm"
    if col_map:
        df = df.rename(columns=col_map)
    for col in expected_cols:
        if col not in df.columns:
            df[col] = None
    if "task_norm" not in df.columns or df["task_norm"].isnull().all():
        if "Task" in df.columns:
            df["task_norm"] = df["Task"].apply(normalize_task)
        else:
            df["task_norm"] = ""
    out = df[expected_cols].copy().where(pd.notnull(df), None)
    out = out.drop_duplicates(subset=["Job Title", "task_norm"], keep="last").reset_index(drop=True)
    rows = out.to_dict(orient="records")
    insert_stmt = text("""
        INSERT INTO all_jobs
        (job_title, task, time_allocation, ai_impact_score, impact_explanation,
         task_transformation, tooling_nature, job_category, Automation_Solution, run_id, jd_hash, task_norm)
        VALUES (:job_title, :task, :time_allocation, :ai_impact_score, :impact_explanation,
                :task_transformation, :tooling_nature, :job_category, :Automation_Solution, :run_id, :jd_hash, :task_norm)
        ON CONFLICT (job_title, task_norm) DO UPDATE SET
            time_allocation = EXCLUDED.time_allocation,
            ai_impact_score = EXCLUDED.ai_impact_score,
            impact_explanation = EXCLUDED.impact_explanation,
            task_transformation = EXCLUDED.task_transformation,
            tooling_nature = EXCLUDED.tooling_nature,
            job_category = EXCLUDED.job_category,
            Automation_Solution = Excluded.Automation_Solution,
            run_id = EXCLUDED.run_id,
            jd_hash = EXCLUDED.jd_hash;
    """)
    stmt_rows = []
    for r in rows:
        stmt_rows.append({
            "job_title": r.get("Job Title"),
            "task": r.get("Task"),
            "time_allocation": r.get("Time allocation %"),
            "ai_impact_score": r.get("AI Impact Score (0–100)"),
            "impact_explanation": r.get("Impact Explanation"),
            "task_transformation": r.get("Task Transformation %"),
            "tooling_nature": r.get("Tooling nature % generic vs specific"),
            "job_category": r.get("Job Category"),
            "Automation_Solution": r.get("Automation Solution"),
            "run_id": r.get("Run ID"),
            "jd_hash": r.get("JD Hash"),
            "task_norm": r.get("task_norm")
        })
    with engine.begin() as conn:
        if stmt_rows:
            conn.execute(insert_stmt, stmt_rows)

def append_synthesis_sql(engine: sa.Engine, syn_rows: list):
    if not syn_rows:
        return
    insert_stmt = text("""
        INSERT INTO synthesis (job_title, synthesis, run_id, jd_hash)
        VALUES (:job_title, :synthesis, :run_id, :jd_hash)
        ON CONFLICT (job_title, jd_hash) DO NOTHING;
    """)
    with engine.begin() as conn:
        conn.execute(insert_stmt, syn_rows)

# -------------------------
# Session-state buffers
# -------------------------
if "new_reports" not in st.session_state:
    st.session_state["new_reports"] = {}
if "new_synthesis" not in st.session_state:
    st.session_state["new_synthesis"] = {}
if "new_jd_text" not in st.session_state:
    st.session_state["new_jd_text"] = {}

# -------------------------
# Sidebar / upload UI
# -------------------------
st.sidebar.header("Upload or Write Job Description(s)")
uploaded_file = st.sidebar.file_uploader("Upload job descriptions (.txt or .csv)", type=["txt", "csv"])
job_text = st.sidebar.text_area("Or paste a single job description here")
st.sidebar.caption(
    "💡 Multiple roles? In .txt, separate with a line containing only `---`. "
    "In .csv, provide one job description per row under a column named 'JobDescription'."
)
generate_clicked_sidebar = st.sidebar.button("🚀 Generate Report", type="primary")

col1, col2 = st.columns([1, 1])
with col1:
    generate_clicked_main = st.button("🚀 Generate Report", type="primary")
with col2:
    powerbi_url = "https://app.powerbi.com/view?r=eyJrIjoiMDFhMGVlOGItOTY5MC00ZTRhLWI5ZTEtNmMwNDQxNTUzNTNmIiwidCI6IjA3NmEzOTkyLTA0ZjgtNDcwMC05ODQ0LTA4YzM3NDc3NzdlZiJ9"
    st.markdown(
        f"""
        <a href="{powerbi_url}" target="_blank">
            <button style="background-color:#0078D4; color:white; padding:0.6em 1.2em; border:none; border-radius:8px; cursor:pointer;">
                📊 Open Dashboard
            </button>
        </a>
        """,
        unsafe_allow_html=True
    )

# -------------------------
# DB connection check (optional)
# -------------------------
engine = None
if DATABASE_URL:
    try:
        engine = get_engine()
        with engine.connect() as c:
            c.execute(sa.text("SELECT 1"))
        st.success("✅ Postgres connection OK (DB commit button will be active).")
        ensure_sql_schema(engine)
    except Exception as e:
        st.warning(f"DB connection failed or not available: {e}. You can still update Excel only.")
        engine = None
else:
    st.info("No DATABASE configured — app will operate in Excel-only mode unless DATABASE is provided.")

# -------------------------
# SYSTEM PROMPT (exact Club Med prompt — updated to handle job title only)
# -------------------------
SYSTEM_PROMPT = """You are GenAI-Job-Impact-Analyst, an expert designed to evaluate how generative AI can transform work at Club Med. 

Your mission
Input: You will receive either
  - a full text job description, OR
  - just a job title (with little or no detail).

If only a job title is given, infer the typical tasks and responsibilities for that role at Club Med or in the hospitality industry, and continue as if a full description was provided.

Output: Produce a table – one line per task – with the following six columns: 
| Task | Job Category | Time allocation % | AI Impact Score (0–100) | Impact Explanation | Task Transformation % | Tooling nature % generic vs specific | Automation Solution |

Task – concise verb-phrase copied, paraphrased, or reasonably inferred from the job title or description. 
Job Category - one of: IT, Marketing, HR, Finance, Operations, Legal, R&D, Customer Service, Other.
Time allocation % – your best estimate of the share of the job’s total time this task takes (sum ≈ 100%). 
AI Impact Score – how strongly Gen-AI could affect the task (0 = no impact, 100 = fully automatable/augmented). 
Impact Explanation – 2–3 sentences justifying the chosen score. Write the Impact Explanation in French.
Task Transformation % – proportion of the task likely to change for the employee (e.g., 70% up-skilling vs 30% pure automation). Always express as two percentages that sum to 100 in the format "XX% up-skilling / YY% automation".
Tooling nature – split the AI tooling you foresee into generic (ChatGPT-like) vs domain-specific (custom models or vertical SaaS). Express as two numbers that sum to 100.
Automation Solution – briefly describe a realistic Gen-AI solution (e.g., "custom GPT-4 powered chatbot", "AI-assisted code generation tool", "AI-driven marketing content generator").

Procedure
A. If a detailed description is given: scan the description and list every distinct, non-trivial activity. 
B. If only a job title is given: generate a reasonable list of 5–10 core tasks typical for the role in hospitality / Club Med.
C. Estimate Time allocation % first – it anchors later scores. Round to nearest 5%.
D. For each activity, consider whether Gen-AI could draft, summarize, translate, ideate, classify, predict or converse, and estimate the effect.
E. Deliver the table in Markdown, then add a short one-paragraph synthesis highlighting the top three automation opportunities and any human-core tasks that should remain manual.

Formatting rules
Use Markdown. Keep lines reasonably wrapped (~80 chars). Round percentages to nearest 5%. Do not invent tasks that are absent when a detailed JD is provided. Never return an empty output — if input is a title only, infer typical tasks and still return a full table + synthesis.
"""

# -------------------------
# GENERATE: parse uploaded/pasted JDs, call Azure OpenAI, buffer outputs
# -------------------------
if generate_clicked_sidebar or generate_clicked_main:
    job_descriptions = []
    if uploaded_file is not None:
        if uploaded_file.name.endswith(".txt"):
            raw_text = uploaded_file.read().decode("utf-8")
            job_descriptions = extract_roles_from_text(raw_text)
        elif uploaded_file.name.endswith(".csv"):
            try:
                df_csv = pd.read_csv(uploaded_file)
            except Exception as e:
                st.error(f"Could not read CSV: {e}")
                st.stop()
            if "JobDescription" in df_csv.columns:
                job_descriptions = df_csv["JobDescription"].dropna().astype(str).tolist()
            else:
                st.error("CSV must contain a column named 'JobDescription'")
                st.stop()
        else:
            st.error("Unsupported file type. Upload .txt or .csv")
            st.stop()
    elif job_text and job_text.strip():
        job_descriptions = [job_text.strip()]
    else:
        st.error("Please upload or paste at least one job description.")
        st.stop()

    with st.spinner("Analyzing job description(s) with Azure OpenAI..."):
        for idx, jd in enumerate(job_descriptions):
            role_name = role_name_from_jobdesc(jd, idx)
            user_prompt = f"Here is the job description or job title:\n\n{jd}"

            try:
                resp = client.chat.completions.create(
                    model=AZURE_OPENAI_DEPLOYMENT,
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ]
                )
                output_text = resp.choices[0].message.content
            except Exception as e:
                st.error(f"OpenAI call failed for {role_name}: {e}")
                output_text = ""

            # Fallback: if the model returned nothing, try an inferred-tasks prompt
            if not output_text or not output_text.strip():
                st.warning(f"No detailed output from model for {role_name}. Using inferred tasks fallback.")
                fallback_prompt = f"Please generate typical tasks for the role '{role_name}' (as used in hospitality/Club Med) and evaluate them following the instructions."
                try:
                    resp = client.chat.completions.create(
                        model=AZURE_OPENAI_DEPLOYMENT,
                        messages=[
                            {"role": "system", "content": SYSTEM_PROMPT},
                            {"role": "user", "content": fallback_prompt},
                        ]
                    )
                    output_text = resp.choices[0].message.content
                except Exception as e:
                    st.error(f"Fallback OpenAI call failed for {role_name}: {e}")
                    output_text = ""

            st.markdown(f"### 📊 Generated Report — **{role_name}**")
            if output_text:
                st.markdown(output_text)
            else:
                st.markdown("_No output from model._")

            table_text, synthesis_text = split_table_and_synthesis(output_text)
            parsed_df = parse_markdown_table(table_text)

            if parsed_df.empty:
                m = re.search(r"(\|.*\|\s*\n\|[-:\s|]+\|\s*\n(?:\|.*\|\s*\n?)*)", output_text, flags=re.DOTALL)
                if m:
                    parsed_df = parse_markdown_table(m.group(0))
                if parsed_df.empty:
                    parsed_df = pd.DataFrame({
                        "Task": [f"[Model output parse failed — see Report]"],
                        "AI Impact Score (0–100)": [None],
                        "Job Category": [None],
                        "Time allocation %": [None],
                        "Impact Explanation": [None],
                        "Task Transformation %": [None],
                        "Tooling nature % generic vs specific": [None],
                        "Automation Solution": [None]
                    })

            if "Job Title" not in parsed_df.columns:
                parsed_df.insert(0, "Job Title", role_name)
            else:
                parsed_df["Job Title"] = parsed_df["Job Title"].replace("", role_name).fillna(role_name)

            canonical_map = {}
            for col in parsed_df.columns:
                lc = col.strip().lower()
                if lc in ["task", "tasks"]:
                    canonical_map[col] = "Task"
                elif "job category" in lc:
                    canonical_map[col] = "Job Category"
                elif "time" in lc and "alloc" in lc:
                    canonical_map[col] = "Time allocation %"
                elif "ai impact" in lc or "impact score" in lc:
                    canonical_map[col] = "AI Impact Score (0–100)"
                elif "impact explanation" in lc or ("explanation" in lc and "impact" in lc):
                    canonical_map[col] = "Impact Explanation"
                elif "task transformation" in lc or "transformation" in lc:
                    canonical_map[col] = "Task Transformation %"
                elif "tooling" in lc:
                    canonical_map[col] = "Tooling nature % generic vs specific"
                elif "automation solution" in lc or ("solution" in lc and "automation" in lc):
                    canonical_map[col] = "Automation Solution"
                elif lc in ["job title", "title"]:
                    canonical_map[col] = "Job Title"
            if canonical_map:
                parsed_df = parsed_df.rename(columns=canonical_map)

            for col in [
                "Task",
                "Job Category",
                "Time allocation %",
                "AI Impact Score (0–100)",
                "Impact Explanation",
                "Task Transformation %",
                "Tooling nature % generic vs specific",
                "Automation Solution"
            ]:
                if col not in parsed_df.columns:
                    parsed_df[col] = None

            run_id = datetime.now().isoformat(timespec="seconds")
            jd_hash = hashlib.sha256(jd.strip().encode("utf-8")).hexdigest()[:12]
            parsed_df["Run ID"] = run_id
            parsed_df["JD Hash"] = jd_hash

            if "Task" in parsed_df.columns:
                parsed_df["task_norm"] = parsed_df["Task"].apply(normalize_task)
                parsed_df = parsed_df.drop_duplicates(subset=["Job Title", "task_norm"], keep="last")
            else:
                parsed_df["task_norm"] = ""

            # Normalize Task Transformation % into "XX% up-skilling / YY% automation"
            def normalize_task_transformation(val):
                if val is None:
                    return None
                s = str(val).strip()
                if "%" in s and ("/" in s or "up" in s.lower()):
                    return s
                m = re.search(r'(\d{1,3})', s)
                if m:
                    num = int(m.group(1))
                    if num < 0: num = 0
                    if num > 100: num = 100
                    other = 100 - num
                    return f"{num}% up-skilling / {other}% automation"
                return None

            if "Task Transformation %" in parsed_df.columns:
                parsed_df["Task Transformation %"] = parsed_df["Task Transformation %"].apply(normalize_task_transformation)

            cols = parsed_df.columns.tolist()
            if "Job Title" in cols:
                cols = ["Job Title"] + [c for c in cols if c != "Job Title"]
                parsed_df = parsed_df[cols]

            st.session_state["new_reports"][role_name] = parsed_df.copy()
            st.session_state["new_synthesis"][role_name] = synthesis_text
            st.session_state["new_jd_text"][role_name] = jd

# -------------------------
# Preview buffered items
# -------------------------
st.divider()
st.subheader("📝 Pending Updates (Buffered; choose how to commit)")

if st.session_state["new_reports"]:
    for role, df in st.session_state["new_reports"].items():
        st.markdown(f"#### {role}")
        try:
            st.dataframe(df, use_container_width=True)
        except Exception:
            st.write(df.head(20))
        syn = st.session_state["new_synthesis"].get(role, "")
        if syn:
            st.markdown("**Synthesis (preview):**")
            st.markdown(syn if len(syn) < 1000 else syn[:1000] + "...")
else:
    st.info("No buffered results. Use 'Generate Report' to parse JDs.")

# -------------------------
# Two separate commit buttons
# -------------------------
st.divider()
st.subheader("Commit options")

col_a, col_b = st.columns(2)

with col_a:
    excel_disabled = not bool(st.session_state["new_reports"])
    if st.button("Update Master Excel (Excel only)", disabled=excel_disabled):
        existing_tasks, existing_syn = load_master_excel()
        try:
            new_tasks = pd.concat(st.session_state["new_reports"].values(), ignore_index=True, sort=False)
        except Exception:
            new_tasks = pd.DataFrame()
            for v in st.session_state["new_reports"].values():
                new_tasks = pd.concat([new_tasks, v], ignore_index=True, sort=False)

        if "Task" in new_tasks.columns:
            new_tasks["task_norm"] = new_tasks["Task"].apply(normalize_task)
        else:
            new_tasks["task_norm"] = ""

        if not existing_tasks.empty and "Task" in existing_tasks.columns:
            existing_tasks["task_norm"] = existing_tasks["Task"].apply(normalize_task)
        else:
            if existing_tasks.empty:
                existing_tasks = pd.DataFrame(columns=new_tasks.columns.tolist())

        if existing_tasks.empty:
            all_tasks = new_tasks.copy()
        else:
            key_cols = ["Job Title", "task_norm"]
            existing_keys = existing_tasks[key_cols].drop_duplicates()
            merged = new_tasks.merge(existing_keys, on=key_cols, how="left", indicator=True)
            to_add = merged[merged["_merge"] == "left_only"].drop(columns=["_merge"])
            all_tasks = pd.concat([existing_tasks, to_add], ignore_index=True, sort=False)

        if "task_norm" in all_tasks.columns:
            cols = [c for c in all_tasks.columns if c != "task_norm"] + ["task_norm"]
            all_tasks = all_tasks[cols]

        new_syn_rows = []
        for role, syn in st.session_state["new_synthesis"].items():
            jd_text = st.session_state["new_jd_text"].get(role, "")
            run_id = datetime.now().isoformat(timespec="seconds")
            jd_hash = hashlib.sha256(jd_text.strip().encode("utf-8")).hexdigest()[:12]
            new_syn_rows.append({
                "Job Title": role,
                "Synthesis": syn,
                "Run ID": run_id,
                "JD Hash": jd_hash
            })
        new_syn_df = pd.DataFrame(new_syn_rows, columns=["Job Title", "Synthesis", "Run ID", "JD Hash"])

        if existing_syn is None or existing_syn.empty:
            all_syn = new_syn_df.copy()
        else:
            all_syn = pd.concat([existing_syn, new_syn_df], ignore_index=True, sort=False)
            if "JD Hash" in all_syn.columns:
                all_syn = all_syn.drop_duplicates(subset=["Job Title", "JD Hash"], keep="last")

        expected_cols_for_excel = [
            "Job Title",
            "Task",
            "Job Category",
            "Time allocation %",
            "AI Impact Score (0–100)",
            "Impact Explanation",
            "Task Transformation %",
            "Tooling nature % generic vs specific",
            "Automation Solution",
            "Run ID",
            "JD Hash"
        ]
        for col in expected_cols_for_excel:
            if col not in all_tasks.columns:
                all_tasks[col] = None
        extras = [c for c in all_tasks.columns if c not in expected_cols_for_excel]
        final_cols = expected_cols_for_excel + extras
        all_tasks = all_tasks[final_cols]

        try:
            buf = write_master_excel(all_tasks, all_syn)
            st.success(f"Master Excel updated: {MASTER_XLSX} (database not modified).")
            st.download_button(
                label="📥 Download Current Master Excel",
                data=buf,
                file_name=MASTER_XLSX,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Failed to write master Excel: {e}")

        st.session_state["new_reports"].clear()
        st.session_state["new_synthesis"].clear()
        st.session_state["new_jd_text"].clear()

with col_b:
    db_disabled = (not bool(st.session_state["new_reports"])) or (engine is None)
    if st.button("Update Database (Postgres only)", disabled=db_disabled):
        if engine is None:
            st.error("No DB engine available. Configure DATABASE to enable DB updates.")
        else:
            try:
                new_tasks = pd.concat(st.session_state["new_reports"].values(), ignore_index=True, sort=False)
            except Exception:
                new_tasks = pd.DataFrame()
                for v in st.session_state["new_reports"].values():
                    new_tasks = pd.concat([new_tasks, v], ignore_index=True, sort=False)

            # Ensure Task Transformation formatting before DB upsert
            def normalize_task_transformation_db(val):
                if val is None:
                    return None
                s = str(val).strip()
                if "%" in s and ("/" in s or "up" in s.lower()):
                    return s
                m = re.search(r'(\d{1,3})', s)
                if m:
                    num = int(m.group(1))
                    if num < 0: num = 0
                    if num > 100: num = 100
                    other = 100 - num
                    return f"{num}% up-skilling / {other}% automation"
                return None

            if "Task Transformation %" in new_tasks.columns:
                new_tasks["Task Transformation %"] = new_tasks["Task Transformation %"].apply(normalize_task_transformation_db)
            else:
                new_tasks["Task Transformation %"] = None

            try:
                upsert_all_jobs_sql(engine, new_tasks)
                # prepare synthesis rows for DB
                syn_rows_for_db = []
                for role, syn in st.session_state["new_synthesis"].items():
                    jd_text = st.session_state["new_jd_text"].get(role, "")
                    syn_rows_for_db.append({
                        "job_title": role,
                        "synthesis": syn,
                        "run_id": datetime.now().isoformat(timespec="seconds"),
                        "jd_hash": hashlib.sha256(jd_text.strip().encode("utf-8")).hexdigest()[:12]
                    })
                if syn_rows_for_db:
                    append_synthesis_sql(engine, syn_rows_for_db)
                st.success("✅ Database updated with buffered rows (Excel not modified).")
            except Exception as e:
                st.error(f"Failed to update Database: {e}")

            # Clear buffers after DB commit to avoid duplicate commits
            st.session_state["new_reports"].clear()
            st.session_state["new_synthesis"].clear()
            st.session_state["new_jd_text"].clear()
